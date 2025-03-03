import sys
import mysql.connector
from PySide6 import QtWidgets, QtCore, QtGui
from PySide6.QtWidgets import QMessageBox, QTableWidgetItem, QFileDialog
from PySide6.QtCore import QFile, QIODevice, Qt
from PySide6.QtUiTools import QUiLoader
import openpyxl
from openpyxl.styles import Font

# Database Configuration (Replace with your actual settings)
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "database": "quan_ly_nong_san"
}

class ReportManagement(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        try: #To run check
            uifile = QFile("report.ui")
            if not uifile.open(QIODevice.ReadOnly):
                print(f"Cannot open {uifile.fileName()}: {uifile.errorString()}")
                sys.exit(-1)

            loader = QUiLoader()
            self.ui = loader.load(uifile, self)
            uifile.close()

            if self.ui is None:
                print(loader.errorString())
                sys.exit(-1)

            self.setCentralWidget(self.ui.centralwidget)
            # Connect the buttons
            self.ui.btnTaoBaoCao.clicked.connect(self.generate_report)

            # Connect export buttons
            self.ui.btnExportDoanhThu.clicked.connect(self.export_doanh_thu)
            self.ui.btnExportSanPham.clicked.connect(self.export_san_pham)
            self.ui.btnExportKhachHang.clicked.connect(self.export_khach_hang)
            self.ui.btnExportKhuyenMai.clicked.connect(self.export_khuyen_mai)

            # Setup table headers
            self.setup_table_headers()
            self.setup_date()

            self.show()
        except Exception as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi: {err}")

    def setup_date(self):
        """Setup the QDateEdit with start and end date"""
        today = QtCore.QDate.currentDate()
        self.ui.dateEditDenNgay.setDate(today) #Set to end date.
        self.ui.dateEditTuNgay.setDate(today) #Set to start date.

    def connect_db(self):
        """Connect to the mySQL database"""
        return mysql.connector.connect(**DB_CONFIG)

    def setup_table_headers(self):
        """Setup the headers for each table widget"""

        # Doanh Thu tab
        self.ui.tableDoanhThu.setColumnCount(3)
        self.ui.tableDoanhThu.setHorizontalHeaderLabels(["Ngày", "Số đơn hàng", "Doanh thu"])
        self.adjust_header(self.ui.tableDoanhThu)

        # San pham ban chay tab
        self.ui.tableSanPhamBanChay.setColumnCount(3)
        self.ui.tableSanPhamBanChay.setHorizontalHeaderLabels(["ID Sản Phẩm", "Tên Sản Phẩm", "Tổng Số Lượng Bán"])
        self.adjust_header(self.ui.tableSanPhamBanChay)

        # Khach hang VIP tab
        self.ui.tableKhachHangVIP.setColumnCount(4)
        self.ui.tableKhachHangVIP.setHorizontalHeaderLabels(["ID Khách Hàng", "Tên Khách Hàng", "Tổng Chi Tiêu", "Số Đơn Hàng"])
        self.adjust_header(self.ui.tableKhachHangVIP)

        # Khuyen mai da dung tab
        self.ui.tableKhuyenMaiDaDung.setColumnCount(3)
        self.ui.tableKhuyenMaiDaDung.setHorizontalHeaderLabels(["ID Khuyến Mãi", "Tên Khuyến Mãi", "Tổng Số Lần Sử Dụng"])
        self.adjust_header(self.ui.tableKhuyenMaiDaDung)

    def adjust_header(self, table_name):
        """Function to config header"""
        header = table_name.horizontalHeader()
        for col in range(table_name.columnCount()):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.ResizeMode.Stretch)

    def generate_report(self):
        """Generate and populate report data in the tables"""
        # Clear previous data in the tables first
        self.clear_tables()

        start_date = self.ui.dateEditTuNgay.date().toString("yyyy-MM-dd")
        end_date = self.ui.dateEditDenNgay.date().toString("yyyy-MM-dd")

        #Fetch Doanh thu data and create the table
        self.load_doanh_thu_data(start_date, end_date)
        #Fetch San pham data and create the table
        self.load_san_pham_data(start_date, end_date)
        #Fetch Khach hang vip data and create the table
        khach_hang_vip_data = self.load_khach_hang_vip_data(start_date, end_date) #Save it to variable
        #Fetch Khuyen mai da dung data and create the table
        self.load_khuyen_mai_data(start_date, end_date)

    def clear_tables(self):
         """Clear the tables first when click the buttons"""
         self.ui.tableDoanhThu.clearContents()
         self.ui.tableDoanhThu.setRowCount(0)
         self.ui.tableSanPhamBanChay.clearContents()
         self.ui.tableSanPhamBanChay.setRowCount(0)
         self.ui.tableKhachHangVIP.clearContents()
         self.ui.tableKhachHangVIP.setRowCount(0)
         self.ui.tableKhuyenMaiDaDung.clearContents()
         self.ui.tableKhuyenMaiDaDung.setRowCount(0)

    def load_doanh_thu_data(self, start_date, end_date):
        """Load Doanh thu data and create the table"""
        conn = self.connect_db()
        cursor = conn.cursor()
        try:
            #Sample SQL
            sql = """SELECT DATE(ngay_tao), COUNT(ma_order), SUM(tong_tien)
                     FROM `order`
                     WHERE ngay_tao BETWEEN %s AND %s
                     GROUP BY DATE(ngay_tao)"""

            # Execute the SQL
            cursor.execute(sql, (start_date, end_date))
            results = cursor.fetchall()

            self.ui.tableDoanhThu.setRowCount(len(results))
            for row, data in enumerate(results):
                self.ui.tableDoanhThu.setItem(row, 0, QTableWidgetItem(str(data[0])))
                self.ui.tableDoanhThu.setItem(row, 1, QTableWidgetItem(str(data[1])))
                self.ui.tableDoanhThu.setItem(row, 2, QTableWidgetItem(str(data[2])))

            #Sum all tong tien
            total_revenue = sum(float(data[2]) for data in results if data[2] is not None)
            self.ui.label_3.setText(f"Tổng doanh thu: {total_revenue:.2f} VNĐ")

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi MySQL: {err}")
        except Exception as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi load Doanh thu: {err}")
        finally:
            cursor.close()
            conn.close()

    def load_san_pham_data(self, start_date, end_date):
        """Load San pham data and create the table"""
        conn = self.connect_db()
        cursor = conn.cursor()
        try:
            #Sample SQL
            sql = """SELECT od.ma_san_pham, sp.ten_san_pham, SUM(od.so_luong) AS total_quantity
                    FROM order_detail od
                    JOIN `order` o ON od.ma_order = o.ma_order
                    JOIN san_pham sp ON od.ma_san_pham = sp.ma_san_pham
                    WHERE o.ngay_tao BETWEEN %s AND %s
                    GROUP BY od.ma_san_pham, sp.ten_san_pham
                    ORDER BY total_quantity DESC
                    LIMIT 10"""
            # Execute the SQL
            cursor.execute(sql, (start_date, end_date))
            results = cursor.fetchall()

            self.ui.tableSanPhamBanChay.setRowCount(len(results))
            for row, data in enumerate(results):
                self.ui.tableSanPhamBanChay.setItem(row, 0, QTableWidgetItem(str(data[0])))
                self.ui.tableSanPhamBanChay.setItem(row, 1, QTableWidgetItem(data[1]))
                self.ui.tableSanPhamBanChay.setItem(row, 2, QTableWidgetItem(str(data[2])))

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi MySQL: {err}")
        except Exception as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi load Sản phẩm bán chạy: {err}")
        finally:
            cursor.close()
            conn.close()

    def load_khach_hang_vip_data(self, start_date, end_date):
        """Load Khach hang vip data and create the table"""
        conn = self.connect_db()
        cursor = conn.cursor()
        try:
            #Sample SQL
            sql = """SELECT o.ma_khach_hang, kh.ten_kh, SUM(o.tong_tien) AS total_spent, COUNT(o.ma_order) AS total_orders
                    FROM `order` o
                    JOIN khach_hang kh ON o.ma_khach_hang = kh.ma_kh
                    WHERE o.ngay_tao BETWEEN %s AND %s
                    GROUP BY o.ma_khach_hang, kh.ten_kh
                    HAVING SUM(o.tong_tien) > 1000000
                    ORDER BY total_spent DESC
                    LIMIT 10"""

            # Execute the SQL
            cursor.execute(sql, (start_date, end_date))
            results = cursor.fetchall()

            # Clear existing data before populating the table
            self.ui.tableKhachHangVIP.setRowCount(len(results))
            for row, data in enumerate(results):
                self.ui.tableKhachHangVIP.setItem(row, 0, QTableWidgetItem(str(data[0])))
                self.ui.tableKhachHangVIP.setItem(row, 1, QTableWidgetItem(data[1]))
                self.ui.tableKhachHangVIP.setItem(row, 2, QTableWidgetItem(str(data[2])))
                self.ui.tableKhachHangVIP.setItem(row, 3, QTableWidgetItem(str(data[3])))
            return results;

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi MySQL: {err}")
        except Exception as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi load Khach hàng: {err}")
        finally:
            cursor.close()
            conn.close()

    def load_khuyen_mai_data(self, start_date, end_date):
        """Load Khuyen mai da dung data and create the table"""
        conn = self.connect_db()
        cursor = conn.cursor()
        try:
            #Sample SQL
            sql = """SELECT o.ma_khuyen_mai, km.ten_ma_km, COUNT(o.ma_khuyen_mai) AS total_used
                    FROM `order` o
                    JOIN khuyen_mai km ON o.ma_khuyen_mai = km.ma_km
                    WHERE o.ngay_tao BETWEEN %s AND %s
                    GROUP BY o.ma_khuyen_mai, km.ten_ma_km
                    ORDER BY total_used DESC
                    LIMIT 10"""
            # Execute the SQL
            cursor.execute(sql, (start_date, end_date))
            results = cursor.fetchall()

            self.ui.tableKhuyenMaiDaDung.setRowCount(len(results))
            for row, data in enumerate(results):
                self.ui.tableKhuyenMaiDaDung.setItem(row, 0, QTableWidgetItem(str(data[0])))
                self.ui.tableKhuyenMaiDaDung.setItem(row, 1, QTableWidgetItem(data[1]))
                self.ui.tableKhuyenMaiDaDung.setItem(row, 2, QTableWidgetItem(str(data[2])))

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi MySQL: {err}")
        except Exception as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi load KM Da Dung: {err}")
        finally:
            cursor.close()
            conn.close()

    def export_doanh_thu(self):
        """Exports the Doanh Thu table data to an Excel file for the selected date range."""
        start_date = self.ui.dateEditTuNgay.date().toString("yyyy-MM-dd")
        end_date = self.ui.dateEditDenNgay.date().toString("yyyy-MM-dd")
        self.export_table_to_excel(self.ui.tableDoanhThu, "Doanh Thu", start_date, end_date)

    def export_san_pham(self):
        """Exports the San Pham Ban Chay table data to an Excel file for the selected date range."""
        start_date = self.ui.dateEditTuNgay.date().toString("yyyy-MM-dd")
        end_date = self.ui.dateEditDenNgay.date().toString("yyyy-MM-dd")
        self.export_table_to_excel(self.ui.tableSanPhamBanChay, "Sản Phẩm Bán Chạy", start_date, end_date)

    def export_khach_hang(self):
        """Exports the Khach Hang VIP table data to an Excel file for the selected date range."""
        start_date = self.ui.dateEditTuNgay.date().toString("yyyy-MM-dd")
        end_date = self.ui.dateEditDenNgay.date().toString("yyyy-MM-dd")
        self.export_table_to_excel(self.ui.tableKhachHangVIP, "Khách Hàng VIP", start_date, end_date)

    def export_khuyen_mai(self):
        """Exports the Khuyen Mai Da Dung table data to an Excel file for the selected date range."""
        start_date = self.ui.dateEditTuNgay.date().toString("yyyy-MM-dd")
        end_date = self.ui.dateEditDenNgay.date().toString("yyyy-MM-dd")
        self.export_table_to_excel(self.ui.tableKhuyenMaiDaDung, "Khuyến Mãi Đã Dùng", start_date, end_date)

    def export_table_to_excel(self, table, tab_name, start_date, end_date):
        """Exports the data from the given QTableWidget to an Excel file, including the date range."""
        conn = self.connect_db()
        cursor = conn.cursor()

        try:
            # Check if there is any data for the selected date range
            if table == self.ui.tableDoanhThu:
                sql = """SELECT 1 FROM `order` WHERE ngay_tao BETWEEN %s AND %s LIMIT 1""" #Check doanh thu
            elif table == self.ui.tableSanPhamBanChay:
                sql = """SELECT 1 FROM order_detail od JOIN `order` o ON od.ma_order = o.ma_order WHERE o.ngay_tao BETWEEN %s AND %s LIMIT 1""" #Check san pham
            elif table == self.ui.tableKhachHangVIP:
                sql = """SELECT 1 FROM `order` o JOIN khach_hang kh ON o.ma_khach_hang = kh.ma_kh WHERE o.ngay_tao BETWEEN %s AND %s LIMIT 1""" #Check khach hang
            elif table == self.ui.tableKhuyenMaiDaDung:
                sql = """SELECT 1 FROM `order` o JOIN khuyen_mai km ON o.ma_khuyen_mai = km.ma_km WHERE o.ngay_tao BETWEEN %s AND %s LIMIT 1""" #Check khuyen mai
            else:
                QMessageBox.critical(self, "Lỗi", "Không xác định được bảng để kiểm tra dữ liệu.")
                return

            cursor.execute(sql, (start_date, end_date))
            has_data = cursor.fetchone()

            if not has_data:
                QMessageBox.information(self, "Thông báo", f"Không có dữ liệu {tab_name} trong khoảng thời gian đã chọn.")
                return

            options = QFileDialog.Options()
            filename, _ = QFileDialog.getSaveFileName(self, f"Lưu file Excel {tab_name}", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

            if filename:
                try:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.title = tab_name

                    # Add date range information
                    sheet['A1'] = f"Báo cáo từ ngày: {start_date} đến ngày: {end_date}"
                    sheet['A1'].font = Font(bold=True)

                    # Write headers starting from the third row
                    headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
                    bold_font = Font(bold=True)
                    for col, header_text in enumerate(headers):
                        cell = sheet.cell(row=3, column=col + 1)  # Start from the third row
                        cell.value = header_text
                        cell.font = bold_font

                    # Write data starting from the fourth row
                    for row in range(table.rowCount()):
                        for col in range(table.columnCount()):
                            item = table.item(row, col)
                            cell = sheet.cell(row=row + 4, column=col + 1)  # Start from the fourth row
                            cell.value = item.text() if item else ""

                    workbook.save(filename)
                    QMessageBox.information(self, "Thông báo", f"Đã xuất dữ liệu {tab_name} thành công vào file {filename}")

                except Exception as e:
                    QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi xuất {tab_name} ra Excel: {e}")

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Lỗi", f"Lỗi MySQL: {err}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi: {e}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = ReportManagement()
    sys.exit(app.exec())