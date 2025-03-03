import sys
import mysql.connector
from PySide6 import QtWidgets, QtCore, QtGui
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QLineEdit, QPushButton,
                             QTableWidget, QTableWidgetItem, QMessageBox, QHeaderView, QDialog, QGroupBox, QFormLayout, QFrame, QDateEdit, QFileDialog)
from PySide6.QtCore import Qt, QFile, QIODevice, QDate
from PySide6.QtUiTools import QUiLoader
import openpyxl
from openpyxl.styles import Font

# Database Configuration
DB_CONFIG = {
    "host": "localhost",  # Replace with your host
    "user": "root",       # Replace with your username
    "password": "",   # Replace with your password
    "database": "quan_ly_nong_san"  # Replace with your database name
}

class OrderDetailWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        uifile = QFile("order_detail.ui")  # Correct the UI file name
        if not uifile.open(QIODevice.ReadOnly):
            print(f"Cannot open {uifile.fileName()}: {uifile.errorString()}")
            sys.exit(-1)

        loader = QUiLoader()
        self.ui = loader.load(uifile, self)
        uifile.close()

        if self.ui is None:
            print(loader.errorString())
            sys.exit(-1)

        self.setCentralWidget(self.ui.centralwidget)
        self.setWindowTitle("Chi Tiết Đơn Hàng")
        self.setGeometry(100, 100, 800, 600)  # Adjust position as needed

        # Connect table click event
        self.ui.tableOrder.itemClicked.connect(self.show_order_details)  # Use tableOrder from UI
        self.ui.btnSearchOrder.clicked.connect(self.search_orders) #Use btnSearchOrder from UI
        self.ui.btnExportExcel.clicked.connect(self.export_to_excel) # Connect export button

        # Add date range search components
        self.date_from = QDateEdit(calendarPopup=True)
        self.date_to = QDateEdit(calendarPopup=True)
        self.date_from.setDate(QDate.currentDate().addDays(-30))  # Default: 30 days ago
        self.date_to.setDate(QDate.currentDate())

        self.btn_search_date = QPushButton("Tìm kiếm theo ngày")
        self.btn_search_date.clicked.connect(self.search_orders_by_date)


        # Add date range search components to UI (inserting at the appropriate place)
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("Từ ngày:"))
        date_layout.addWidget(self.date_from)
        date_layout.addWidget(QLabel("Đến ngày:"))
        date_layout.addWidget(self.date_to)
        date_layout.addWidget(self.btn_search_date)

        self.ui.verticalLayout.insertLayout(2, date_layout)  # Insert after lblOrderDetails

        # Load Initial Data
        self.load_orders()

    def connect_db(self):
        """Connect to the MySQL database."""
        return mysql.connector.connect(**DB_CONFIG)

    def load_orders(self):
        """Load all order and order detail IDs and creation dates."""
        conn = self.connect_db()
        cursor = conn.cursor()

        try:
            # Load Order IDs and Creation Dates
            sql_orders = """
                SELECT DISTINCT od.ma_order, o.ngay_tao
                FROM order_detail od
                JOIN `order` o ON od.ma_order = o.ma_order
                ORDER BY od.ma_order
            """
            cursor.execute(sql_orders)
            orders = cursor.fetchall()

            self.ui.tableOrder.setRowCount(len(orders))  # Use ui table
            self.ui.tableOrder.setColumnCount(2)
            self.ui.tableOrder.setHorizontalHeaderLabels(["Mã Đơn Hàng", "Ngày Tạo"])


            for row, order in enumerate(orders):
                self.ui.tableOrder.setItem(row, 0, QTableWidgetItem(str(order[0]))) # Ma Order
                date_value = order[1] #This will now retrieve the date from the result
                if date_value:
                    date_str = date_value.strftime('%Y-%m-%d %H:%M:%S')  # Format date as YYYY-MM-DD HH:MM:SS (more precise)
                    self.ui.tableOrder.setItem(row, 1, QTableWidgetItem(date_str))  # Ngay Tao
                else:
                    self.ui.tableOrder.setItem(row,1, QTableWidgetItem(""))  # Handle potential NULL values



        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Database Error", f"Error loading orders: {err}")
        finally:
            cursor.close()
            conn.close()


    def search_orders(self):
        """Search orders based on ma_order and also load the corresponding order details."""
        search_term = self.ui.txtSearchOrder.text().strip() #Use ui textedit

        conn = self.connect_db()
        cursor = conn.cursor()

        try:
            if search_term:
                # Search for matching order IDs
                sql_orders = """
                    SELECT DISTINCT od.ma_order, o.ngay_tao
                    FROM order_detail od
                    JOIN `order` o ON od.ma_order = o.ma_order
                    WHERE od.ma_order LIKE %s
                    ORDER BY od.ma_order
                """
                cursor.execute(sql_orders, (f"%{search_term}%",))
                orders = cursor.fetchall()

                self.ui.tableOrder.setRowCount(len(orders)) #Use ui table
                self.ui.tableOrder.setColumnCount(2)
                self.ui.tableOrder.setHorizontalHeaderLabels(["Mã Đơn Hàng", "Ngày Tạo"])

                for row, order in enumerate(orders):
                     self.ui.tableOrder.setItem(row, 0, QTableWidgetItem(str(order[0]))) # Ma Order
                     date_value = order[1] #This will now retrieve the date from the result
                     if date_value:
                        date_str = date_value.strftime('%Y-%m-%d %H:%M:%S')  # Format date as YYYY-MM-DD HH:MM:SS (more precise)
                        self.ui.tableOrder.setItem(row, 1, QTableWidgetItem(date_str))  # Ngay Tao
                     else:
                        self.ui.tableOrder.setItem(row,1, QTableWidgetItem(""))  # Handle potential NULL values
            else:
                self.load_orders()  # If search term is empty, reload all orders


        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Database Error", f"Error searching orders: {err}")

        finally:
            cursor.close()
            conn.close()

    def search_orders_by_date(self):
        """Search orders within a specified date range."""
        start_date = self.date_from.date().toString(Qt.ISODate)
        end_date = self.date_to.date().toString(Qt.ISODate)

        conn = self.connect_db()
        cursor = conn.cursor()

        try:
            # Adjust your SQL query to filter by date.  You'll need to know
            # the column in your table that stores the date.  I'm assuming it's called 'ngay_tao'.
            sql_orders = """
                SELECT DISTINCT od.ma_order, o.ngay_tao
                FROM order_detail od
                JOIN `order` o ON od.ma_order = o.ma_order
                WHERE DATE(o.ngay_tao) BETWEEN %s AND %s
                ORDER BY od.ma_order
            """
            cursor.execute(sql_orders, (start_date, end_date))
            orders = cursor.fetchall()

            self.ui.tableOrder.setRowCount(len(orders))
            self.ui.tableOrder.setColumnCount(2)
            self.ui.tableOrder.setHorizontalHeaderLabels(["Mã Đơn Hàng", "Ngày Tạo"])

            for row, order in enumerate(orders):
                 self.ui.tableOrder.setItem(row, 0, QTableWidgetItem(str(order[0]))) # Ma Order
                 date_value = order[1] #This will now retrieve the date from the result
                 if date_value:
                    date_str = date_value.strftime('%Y-%m-%d %H:%M:%S')  # Format date as YYYY-MM-DD HH:MM:SS (more precise)
                    self.ui.tableOrder.setItem(row, 1, QTableWidgetItem(date_str))  # Ngay Tao
                 else:
                    self.ui.tableOrder.setItem(row,1, QTableWidgetItem(""))  # Handle potential NULL values


        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Database Error", f"Error searching orders by date: {err}")
        finally:
            cursor.close()
            conn.close()

    def show_order_details(self, item):
        """Display detailed information about the selected order in a new dialog."""
        order_id = item.text()  #Get the order ID from the clicked table item
        self.order_details_dialog = OrderDetailsDialog(order_id) #Creating the Dialog
        self.order_details_dialog.show()

    def export_to_excel(self):
        """Exports order data to an Excel file.
        If a search term is present, exports the *details* for orders matching that
        search term. If no search term is present, *and* dates have been selected and
        "Search By Date" has been clicked, exports all order details for that date range.
        Otherwise, warns the user to select a date range.
        """
        options = QFileDialog.Options()
        filename, _ = QFileDialog.getSaveFileName(self, "Lưu file Excel", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

        if filename:
            search_term = self.ui.txtSearchOrder.text().strip()
            start_date = self.date_from.date().toString(Qt.ISODate)
            end_date = self.date_to.date().toString(Qt.ISODate)

            if search_term:
                # Export order details for the matching order(s)
                self.export_order_details_by_search_term(filename, search_term)
            elif start_date and end_date:
                # Export all order details for the selected date range
                self.export_all_order_details_by_date(filename, start_date, end_date)
            else:
                QMessageBox.warning(self, "Thông báo", "Vui lòng chọn khoảng thời gian trước khi xuất Excel (nếu không có mã đơn hàng cụ thể).")


    def export_order_details_by_search_term(self, filename, search_term):
        """Export the *details* of orders matching the search term."""
        conn = self.connect_db()
        cursor = conn.cursor()

        try:
            #SQL query to get the order details
            sql = """
                SELECT sp.ma_san_pham, sp.ten_san_pham, od.so_luong, od.gia_ban,
                od.so_luong * od.gia_ban AS tong_tien, kh.ten_kh, o.ngay_tao
                FROM order_detail od
                JOIN san_pham sp ON od.ma_san_pham = sp.ma_san_pham
                LEFT JOIN `order` o ON od.ma_order = o.ma_order
                LEFT JOIN khach_hang kh ON o.ma_khach_hang = kh.ma_kh
                WHERE od.ma_order LIKE %s
            """

            cursor.execute(sql, (f"%{search_term}%",))
            order_details = cursor.fetchall()

            if not order_details:
                QMessageBox.information(self, "Thông báo", "Không có chi tiết đơn hàng cho mã đơn hàng này.")
                return

            #Create a new Excel workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            #Write the headers to the first row, making them bold
            headers = ["Mã SP", "Tên SP", "Số Lượng", "Giá Bán", "Tổng Tiền", "Tên KH", "Ngày Tạo"]
            bold_font = Font(bold=True)
            for col, header_text in enumerate(headers):
                cell = sheet.cell(row=1, column=col + 1)
                cell.value = header_text
                cell.font = bold_font

            #Write the data to the subsequent rows
            for row, detail in enumerate(order_details, start=2):
                sheet.cell(row=row, column=1).value = detail[0]  #Ma SP
                sheet.cell(row=row, column=2).value = detail[1]  #Ten SP
                sheet.cell(row=row, column=3).value = detail[2]  #So Luong
                sheet.cell(row=row, column=4).value = detail[3]  #Gia Ban
                sheet.cell(row=row, column=5).value = detail[4]  #Tong Tien
                sheet.cell(row=row, column=6).value = detail[5] if detail[5] else ""  #Ten KH
                sheet.cell(row=row, column=7).value = detail[6].strftime('%Y-%m-%d %H:%M:%S') if detail[6] else "" # Ngay Tao

             #Save the workbook
            workbook.save(filename)
            QMessageBox.information(self, "Thông báo", f"Đã xuất dữ liệu thành công vào file {filename}")

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Lỗi cơ sở dữ liệu", f"Lỗi xuất Excel: {err}")

        except Exception as e:
             QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi: {e}")

        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    def export_all_order_details_by_date(self, filename, start_date, end_date):
        """Exports *all* order details within a specified date range."""
        conn = self.connect_db()
        cursor = conn.cursor()

        try:
            sql = """
                SELECT od.ma_order, sp.ma_san_pham, sp.ten_san_pham, od.so_luong, od.gia_ban,
                od.so_luong * od.gia_ban AS tong_tien, kh.ten_kh, o.ngay_tao
                FROM order_detail od
                JOIN san_pham sp ON od.ma_san_pham = sp.ma_san_pham
                LEFT JOIN `order` o ON od.ma_order = o.ma_order
                LEFT JOIN khach_hang kh ON o.ma_khach_hang = kh.ma_kh
                WHERE DATE(o.ngay_tao) BETWEEN %s AND %s
                ORDER BY od.ma_order
            """
            cursor.execute(sql, (start_date, end_date))
            all_order_details = cursor.fetchall()

            if not all_order_details:
                QMessageBox.information(self, "Thông báo", "Không có chi tiết đơn hàng nào trong khoảng thời gian này.")
                return

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Header row
            headers = ["Mã SP", "Tên SP", "Số Lượng", "Giá Bán", "Tổng Tiền", "Tên KH", "Ngày Tạo"]
            bold_font = Font(bold=True)
            for col, header_text in enumerate(headers):
                cell = sheet.cell(row=1, column=col + 2)  # Start from column 2, leaving space for order ID
                cell.value = header_text
                cell.font = bold_font

            # Write data, grouped by order ID, with order ID as a heading
            current_row = 2
            current_order_id = None

            for detail in all_order_details:
                order_id = detail[0]

                if order_id != current_order_id:
                    # New order ID - add a heading
                    current_order_id = order_id
                    sheet.cell(row=current_row, column=1).value = f"Mã Đơn Hàng: {order_id}"
                    sheet.cell(row=current_row, column=1).font = bold_font
                    current_row += 1  # Move to the next row after writing order ID

                # Write the order details
                sheet.cell(row=current_row, column=2).value = detail[1]  # Ma SP
                sheet.cell(row=current_row, column=3).value = detail[2]  # Ten SP
                sheet.cell(row=current_row, column=4).value = detail[3]  # So Luong
                sheet.cell(row=current_row, column=5).value = detail[4]  # Gia Ban
                sheet.cell(row=current_row, column=6).value = detail[5]  # Tong Tien
                sheet.cell(row=current_row, column=7).value = detail[6] if detail[6] else ""  # Ten KH
                sheet.cell(row=current_row, column=8).value = detail[7].strftime('%Y-%m-%d %H:%M:%S') if detail[7] else ""  # Ngay Tao

                current_row += 1  # Increment the row number

            workbook.save(filename)
            QMessageBox.information(self, "Thông báo", f"Đã xuất dữ liệu thành công vào file {filename}")

        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Lỗi cơ sở dữ liệu", f"Lỗi xuất Excel: {err}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi: {e}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

class OrderDetailsDialog(QDialog):  #Separate class for the details dialog
    def __init__(self, order_id):
        super().__init__()
        self.order_id = order_id #Storing order ID
        self.setWindowTitle(f"Chi Tiết Đơn Hàng - Mã: {order_id}")
        self.setGeometry(200, 200, 800, 600)

        self.main_layout = QVBoxLayout(self)  # Use a main layout for the dialog

        # Table to Display Order Details
        self.order_details_table = QTableWidget()
        self.order_details_table.setColumnCount(6)  # Define number of columns
        self.order_details_table.setHorizontalHeaderLabels(["Mã SP", "Tên SP", "Số Lượng", "Giá Bán", "Tổng Tiền", "Tên KH"])

        header = self.order_details_table.horizontalHeader()
        for col in range(self.order_details_table.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)

        self.main_layout.addWidget(self.order_details_table)

        self.load_order_details()

    def connect_db(self):
        """Connect to the MySQL database."""
        return mysql.connector.connect(**DB_CONFIG)

    def load_order_details(self):
        """Load detailed information for the specified order from the database."""
        conn = self.connect_db()
        cursor = conn.cursor()

        try:
            sql = """
                SELECT sp.ma_san_pham, sp.ten_san_pham, od.so_luong, od.gia_ban,
                od.so_luong * od.gia_ban AS tong_tien, kh.ten_kh, o.ngay_tao
                FROM order_detail od
                JOIN san_pham sp ON od.ma_san_pham = sp.ma_san_pham
                LEFT JOIN `order` o ON od.ma_order = o.ma_order  -- Join order_detail and order
                LEFT JOIN khach_hang kh ON o.ma_khach_hang = kh.ma_kh   -- JOIN order and customer
                WHERE od.ma_order = %s
            """
            print(f"SQL Query: {sql % (self.order_id,)}")  # Print the query with the parameter
            cursor.execute(sql, (self.order_id,))
            order_details = cursor.fetchall()

            print(f"Number of order details found: {len(order_details)}")  # Check results

            self.order_details_table.setRowCount(len(order_details))
            # self.order_details_table.setColumnCount(7)  # Update column count  <--  REMOVE THIS LINE!
            # self.order_details_table.setHorizontalHeaderLabels(["Mã SP", "Tên SP", "Số Lượng", "Giá Bán", "Tổng Tiền", "Tên KH", "Ngày Tạo"]) #Update header <--  REMOVE THIS LINE!

            for row, detail in enumerate(order_details):
                # Populate the table with the retrieved data
                self.order_details_table.setItem(row, 0, QTableWidgetItem(str(detail[0])))  # Ma SP
                self.order_details_table.setItem(row, 1, QTableWidgetItem(detail[1]))  # Ten SP
                self.order_details_table.setItem(row, 2, QTableWidgetItem(str(detail[2])))  # So luong mua
                self.order_details_table.setItem(row, 3, QTableWidgetItem(str(detail[3])))  # Gia ban
                self.order_details_table.setItem(row, 4, QTableWidgetItem(str(detail[4])))  # Tong tien
                self.order_details_table.setItem(row, 5, QTableWidgetItem(detail[5] or ""))  # Ten khach hang. Handles NULL values
               # date_value = detail[6]  # Get date value from the query result <--  REMOVE THIS LINE!
               # if date_value:        <--  REMOVE THIS LINE!
               #     date_str = date_value.strftime('%Y-%m-%d %H:%M:%S')  # Format date as YYYY-MM-DD HH:MM:SS  <--  REMOVE THIS LINE!
               #     self.order_details_table.setItem(row, 6, QTableWidgetItem(date_str))  # Ngay Tao   <--  REMOVE THIS LINE!
               # else:                 <--  REMOVE THIS LINE!
               #     self.order_details_table.setItem(row, 6, QTableWidgetItem(""))    <--  REMOVE THIS LINE!
        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Database Error", f"Error loading order details: {err}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = OrderDetailWindow()
    window.show()
    sys.exit(app.exec())